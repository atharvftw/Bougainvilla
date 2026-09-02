#!/usr/bin/env python3
"""
Build September-2026-Content-Calendar.xlsx for Bougainvilla Resorts.

Mirrors the structure of the August 2026 calendar (Sheet 1 grid + Sheet 2
designer brief) and adds the stories, highlights, captions, inspo and
production sheets planned in docs/september-2026-plan.md.

No credentials from the source Google Sheets are reproduced in this file.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "deliverables/September-2026-Content-Calendar.xlsx"

# ---------------------------------------------------------------- palette
# Values carried from the August Designer Brief.
CREAM       = "F2EBE1"
WARM_WHITE  = "FBF9F6"
SAND        = "C8A882"
BRAND_BROWN = "6B4F3A"
TERRACOTTA  = "B5714F"

# Tints used for the calendar grid's type colour-coding. Light enough that
# brand-brown body text stays legible on top.
FILL_REEL     = "E8D5CC"
FILL_CAROUSEL = "EFE3D2"
FILL_STATIC   = "DED3C6"
FILL_STORY    = CREAM
FILL_FESTIVAL = "C9B49A"
FILL_QUIET    = WARM_WHITE
FILL_HEADER   = BRAND_BROWN
FILL_SUBHEAD  = SAND

FONT = "Arial"

def f(size=10, bold=False, color=BRAND_BROWN, italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)

def fill(hexcolor):
    return PatternFill("solid", fgColor=hexcolor)

def align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

THIN = Side(style="thin", color=SAND)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_cell(ws, ref, value=None, font=None, fillc=None, alignment=None,
               border=True):
    c = ws[ref]
    if value is not None:
        c.value = value
    c.font = font or f()
    if fillc:
        c.fill = fill(fillc)
    c.alignment = alignment or align()
    if border:
        c.border = BOX
    return c

def banner(ws, row, text, span=7, size=11, bg=FILL_HEADER, fg=CREAM, bold=True):
    """Full-width merged banner row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=size, bold=bold, color=fg)
    c.fill = fill(bg)
    c.alignment = align("left", "center")
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).border = BOX
    return c

# ------------------------------------------------------------- inspo link table
# Supplied by the client. NOT REVIEWED — instagram.com is blocked by this
# session's network egress proxy, so none of these could be opened here.
LINKS = {
    "REEL 1 · All Of It":                    "https://www.instagram.com/p/DadUztOMe56/",
    "STATIC 1 · The Lake Is The View":       "https://www.instagram.com/p/DcQEepBDakg/?img_index=2",
    "CAROUSEL 1 · What A Weekday Gets You":  "https://www.instagram.com/p/DbsuoOVjPzV/",
    "REEL 2 · The Pool, Day To Night":       "https://www.instagram.com/p/DceNoWbyQnA/",
    "CAROUSEL 2 · Seven Things Inside":      "https://www.instagram.com/p/DcVqkH6CP9E/",
    "STATIC 2 · Ganpati Greeting":           "https://www.instagram.com/p/DbA4f-sNjtf/",
    "CAROUSEL 3 · A Guide To Karjat":        "https://www.instagram.com/p/Db71wKNAOQr/",
    "REEL 3 · After The City":               "https://www.instagram.com/p/DbvH3GeJSR0/",
    "CAROUSEL 4 · Who It's For":             "https://www.instagram.com/p/DcX2SH-iByn/",
    "REEL 4 · October Is Clear":             "https://www.instagram.com/p/DcYxp75p8GM/",
    "REEL 5 · Book Now":                     "https://www.instagram.com/p/DcOTWy6ML1t/?igsi=MWtjOHB0NnIybzgwYw==",
    "FONT REFERENCE":                        "https://www.instagram.com/p/Dblrpj-E_Lf/?img_index=1&igsi=cWd0amt3MXdqZnAw",
}
ACCOUNT   = "https://www.instagram.com/bougainvillaresorts/"
AUGUST_CC = "https://docs.google.com/spreadsheets/d/11LoupuCnND-cbPLo_sRouKiM7wiTdfpYL_T730yC8_k/edit"
SOURCE_CC = "https://docs.google.com/spreadsheets/d/16MkXCAyTYa9LaKJd0rAt-JwzQyvtple8v7dhTFcl3SE/edit"

GRID_LINKS = {
    5:  ("FONT REFERENCE", "Font reference — pick ONE face"),
    7:  ("REEL 1 · All Of It", "Reel 1 inspo"),
    8:  ("STATIC 1 · The Lake Is The View", "Static 1 inspo"),
    9:  ("CAROUSEL 1 · What A Weekday Gets You", "Carousel 1 inspo"),
    10: ("REEL 2 · The Pool, Day To Night", "Reel 2 inspo"),
    11: ("CAROUSEL 2 · Seven Things Inside", "Carousel 2 inspo"),
    14: ("STATIC 2 · Ganpati Greeting", "Static 2 inspo"),
    16: ("CAROUSEL 3 · A Guide To Karjat", "Carousel 3 inspo"),
    21: ("REEL 3 · After The City", "Reel 3 inspo"),
    23: ("CAROUSEL 4 · Who It's For", "Carousel 4 inspo"),
    28: ("REEL 4 · October Is Clear", "Reel 4 inspo"),
    30: ("REEL 5 · Book Now", "Reel 5 inspo"),
}

wb = Workbook()

# ==================================================================== SHEET 1
ws = wb.active
ws.title = "September 2026"

banner(ws, 1, "SEPTEMBER 2026  ·  BOUGAINVILLA RESORTS  ·  CONTENT CALENDAR", size=13)
ws.row_dimensions[1].height = 28
banner(ws, 2,
       "RUNS MON 7 SEP → WED 30 SEP  ·  10% off Mon–Thu  ·  rates on the OFFERS highlight only  ·  "
       "WhatsApp only  ·  Ganpati Sep 14–24  ·  Pitru Paksha from Sep 26",
       size=10, bg=FILL_SUBHEAD, fg=BRAND_BROWN, bold=False)
ws.row_dimensions[2].height = 22

DAYS = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
for i, d in enumerate(DAYS, start=1):
    c = ws.cell(row=3, column=i, value=d)
    c.font = f(10, bold=True, color=CREAM)
    c.fill = fill(BRAND_BROWN)
    c.alignment = align("center", "center")
    c.border = BOX
ws.row_dimensions[3].height = 20

# (date_label, body_text, fill) per day; None = no day in this cell
R = "reel"; C = "car"; S = "stat"; T = "story"; F_ = "fest"; Q = "quiet"
KIND_FILL = {R: FILL_REEL, C: FILL_CAROUSEL, S: FILL_STATIC,
             T: FILL_STORY, F_: FILL_FESTIVAL, Q: FILL_QUIET}

WEEKS = [
    # ---- production week : Sep 1-6, nothing posts
    [
        (None, "", None),
        ("1", "NO POST — production week.\n\nThe calendar starts Mon Sep 7 and finishes Wed "
              "Sep 30. Nothing publishes before Sep 7.", Q),
        ("2", "SHOOT — everything for week 1.\n\nReel 1, Static 1, Carousel 1, Reel 2, "
              "Carousel 2 and Stories 1–10 all publish between Sep 7 and Sep 13. They must be "
              "shot and cut now.", Q),
        ("3", "SHOOT / EDIT", Q),
        ("4", "SHOOT / EDIT", Q),
        ("5", "DESIGN — Carousel 1 and Static 1 to the designer.", Q),
        ("6", "APPROVALS due end of day.\n\nWeek 1 carries 5 of the month's 11 feed pieces. "
              "Nothing in it can slip.", Q),
    ],
    # ---- week 1 : Sep 7-13   THE ONLY SELLING WEEK
    [
        ("7", "REEL 1 · 9:30 PM\nALL OF IT\n\nThe whole house in one take — exterior, living, "
              "stairs, pool, lake. Opens the month.\n\nCTA: WhatsApp\n\n────────\n"
              "STORY 1 · 9:30 AM\nVILLA — overall tour", R),
        ("8", "STATIC 1 · 9:30 AM\nTHE LAKE IS THE VIEW\n\nLake-facing hero. The asset August "
              "never used.\n\nNumber + WhatsApp icon, brand brown.\n\n────────\n"
              "STORY 2 · 9:30 PM\nROOMS — all four bedrooms", S),
        ("9", "CAROUSEL 1 · 9:30 AM\nWHAT A WEEKDAY GETS YOU\n\nTHE OFFER CAROUSEL. This is "
              "the only selling week in the month — it has to land here.\n\n7 slides. Final "
              "slide = WhatsApp card. No rupee figures.\n\n────────\nSTORY 3 · 9:30 PM\nPOOL", C),
        ("10", "REEL 2 · 9:30 PM\nTHE POOL, DAY TO NIGHT\n\nMorning → sunset → night → lake."
               "\n\nThe 2 PM weekday shot with nobody in it is the one that sells.\n\n"
               "────────\nSTORY 4 · 9:30 AM\nLAKE VIEW", R),
        ("11", "CAROUSEL 2 · 9:30 AM\nSEVEN THINGS INSIDE\n\nPool · lake · theatre · gym · "
               "games · library · garden.\n\nCTA: Save + WhatsApp\n\n────────\n"
               "STORIES 5–6 · 9:30 PM\nAMENITIES · GAMES [Poll]", C),
        ("12", "STORIES 7–8 · 9:30 AM\n\n7 · THEATRE\n8 · KITCHEN\n\n>> SEEDS: EXPERIENCES", T),
        ("13", "STORIES 9–10 · 9:30 AM\n\n9 · OUTDOOR\n10 · CAPACITY [Poll]\n\n"
               "GANPATI EVE — no feed post.\n\n>> BUILD HIGHLIGHTS:\n"
               "VILLA · ROOMS · POOL · EXPERIENCES", T),
    ],
    # ---- week 2 : Sep 14-20   GANPATI
    [
        ("14", "GANESH CHATURTHI [CONFIRM DATE]\n\nSTATIC 2 · 9:00 AM\nGANPATI GREETING\n\n"
               "Greeting only. NO offer, NO CTA, NO discount, NO clip-art, NO murti imagery.\n\n"
               "A diya or marigold detail and the villa. That is all.", F_),
        ("15", "Silent.\n\nDay 2 of the festival. Nothing posted.", Q),
        ("16", "CAROUSEL 3 · 9:30 AM\nA GUIDE TO KARJAT\n\nScenery, nearby spots, the drive, "
               "escape-the-city.\n\nBUILT FOR SAVES — saves compound in a high-attention week. "
               "Make slide 1 look like a guide, not an ad.\n\nCTA: Save + WhatsApp", C),
        ("17", "STORY 11 · 9:30 AM\nLOCATION\n\nKarjat + Google Maps. [Location sticker]\n\n"
               ">> SEEDS: KARJAT", T),
        ("18", "STORY 12 · 9:30 PM\nHOW TO REACH\n\nMumbai and Pune routes, landmarks, drive "
               "time. [Question box]\n\n>> SEEDS: KARJAT", T),
        ("19", "STORY 13 · 9:30 AM\nPET FRIENDLY\n\nPolicy + which spaces. [Question box]", T),
        ("20", "Quiet. Gauri and visarjan for many households.\n\n>> BUILD HIGHLIGHT: KARJAT", Q),
    ],
    # ---- week 3 : Sep 21-27   RE-ENTRY
    [
        ("21", "REEL 3 · 9:30 PM\nAFTER THE CITY\n\nTen days of the city, now this. Groups, "
               "friends, celebrations.\n\nFirst push since Sep 11.\n\nCTA: WhatsApp\n\n"
               "────────\nSTORY 14 · 9:30 AM\nSTAY", R),
        ("22", "STORY 15 · 9:30 AM\nCELEBRATIONS\n\nBirthdays, anniversaries, gatherings.\n\n"
               ">> SEEDS: GROUPS", T),
        ("23", "CAROUSEL 4 · 9:30 AM\nWHO IT'S FOR\n\nFriends · families · celebrations · "
               "capacity · private villa.\n\nThe whole villa goes to one group. Always.\n\n"
               "CTA: WhatsApp\n\n────────\nSTORY 16 · 9:30 PM\nREVIEWS", C),
        ("24", "ANANT CHATURDASHI · VISARJAN [CONFIRM DATE]\n\nSilent. No post.", F_),
        ("25", "STORY 17 · 9:30 AM\nCHECK-IN\n\nProcess and timings.", T),
        ("26", "PITRU PAKSHA BEGINS [CONFIRM]\n\nSTORY 18 · 9:30 AM\nRULES\n\n"
               "No offer content from here to month end.", F_),
        ("27", "Quiet.\n\n>> BUILD HIGHLIGHT: GROUPS", Q),
    ],
    # ---- week 4 : Sep 28-30   CLOSE
    [
        ("28", "REEL 4 · 9:30 PM\nOCTOBER IS CLEAR\n\nPost-monsoon — the green without the "
               "rain. Pre-sells Navratri and Diwali.\n\nNo September offer anywhere in this."
               "\n\n────────\nSTORY 19 · 9:30 AM\nFAQs [Question box]", R),
        ("29", "STORY 20 · 9:30 AM\nBOOK NOW\n\nWhatsApp / call / DM. [Link sticker]\n\n"
               ">> SEEDS: OFFERS", T),
        ("30", "REEL 5 · 9:30 PM\nBOOK NOW\n\nCloses the month on conversion, pointed at "
               "October.\n\n>> BUILD HIGHLIGHT: OFFERS\n(built last, pinned first)\n\n"
               ">> SET TRAY ORDER — add one frame to each highlight in REVERSE order today.", R),
        (None, "", None),
        (None, "", None),
        (None, "", None),
        (None, "", None),
    ],
]

row = 4
for week in WEEKS:
    # date row
    for i, (label, _body, kind) in enumerate(week, start=1):
        c = ws.cell(row=row, column=i, value=label if label else "")
        c.font = f(11, bold=True)
        c.fill = fill(FILL_FESTIVAL if kind == F_ else FILL_SUBHEAD if label else FILL_QUIET)
        c.alignment = align("left", "center")
        c.border = BOX
    ws.row_dimensions[row].height = 20
    # body row
    for i, (_label, body, kind) in enumerate(week, start=1):
        c = ws.cell(row=row + 1, column=i, value=body)
        c.font = f(8.5)
        c.fill = fill(KIND_FILL.get(kind, FILL_QUIET))
        c.alignment = align("left", "top")
        c.border = BOX
    ws.row_dimensions[row + 1].height = 172
    row += 2

# ---- stamp the inspo links into the grid cells and make them clickable
DATE_POS = {}
body_rows = [5, 7, 9, 11, 13]
for bi, week in enumerate(WEEKS):
    for ci, (lbl, _b, _k) in enumerate(week, start=1):
        if lbl and lbl.isdigit():
            DATE_POS[int(lbl)] = (body_rows[bi], ci)

for day, (piece, cap) in GRID_LINKS.items():
    br, col = DATE_POS[day]
    cell = ws.cell(row=br, column=col)
    url = LINKS[piece]
    cell.value = (cell.value or "") + "\n\n▸ " + cap + ":\n" + url
    cell.hyperlink = url
    cell.font = f(8.5, color="1155CC")

# ---- legend + notes
row += 1
banner(ws, row, "LEGEND"); row += 1
legend = [("Reel", FILL_REEL), ("Carousel", FILL_CAROUSEL), ("Static", FILL_STATIC),
          ("Story", FILL_STORY), ("Quiet / no post", FILL_QUIET),
          ("Festival — confirm date", FILL_FESTIVAL)]
for i, (label, colr) in enumerate(legend, start=1):
    style_cell(ws, f"{get_column_letter(i)}{row}", label, f(9, bold=True), colr,
               align("center", "center"))
ws.row_dimensions[row].height = 20
row += 2

banner(ws, row, "POSTING TIMES — from Instagram Insights, unchanged from August"); row += 1
banner(ws, row, "Reels 9:30 PM (peak, 143)  ·  Carousels + Statics 9:30 AM "
                "(morning scroll, 158–159)  ·  NEVER 12 PM–4 PM (trough, 25–59)",
       bg=WARM_WHITE, fg=BRAND_BROWN, bold=False); row += 2

banner(ws, row, "MONTH RULES — SEPTEMBER"); row += 1
RULES = [
    "CARRIED FROM AUGUST — No rupee figures in the feed. No 'limited time', no countdowns, "
    "no false scarcity. Every offer ends in WhatsApp, never link in bio.",
    "CARRIED FROM AUGUST — WhatsApp icon + number on every static: bottom third, brand brown, "
    "not green. Every carousel ends on a full WhatsApp slide, not a corner overlay.",
    "NEW — Prices appear on the OFFERS highlight frames and NOWHERE ELSE. This is a scoped "
    "exception to the August no-rates rule, not permission to put rates on statics or in feed.",
    "NEW — Shravan has ended, so non-veg and bar imagery are permitted again. Do not use them "
    "from Sep 14 onward: much of this audience is vegetarian through Ganpati and strictly so "
    "through Pitru Paksha. Safe window is Sep 7–13 only.",
    "NEW — The month posts Sep 7–30. Sep 1–6 is production time, nothing publishes. No offer "
    "content at all Sep 14–24 (Ganpati) or Sep 26–30 (Pitru Paksha) — which leaves Sep 7–13 as "
    "the ONLY selling week. It carries 5 of the 11 feed pieces. Nothing in it can slip.",
    "NEW — No Ganpati murti imagery, no idol photography, no visarjan footage. A greeting, a "
    "diya, a marigold detail, the villa. Nothing devotional used as a marketing frame.",
    "NEW — The lake is a headline asset this month. It appears nowhere in August. It gets a "
    "static, a reel beat, a story and a slot in three highlights.",
]
for r in RULES:
    banner(ws, row, r, bg=WARM_WHITE, fg=BRAND_BROWN, bold=False, size=9)
    ws.row_dimensions[row].height = 26
    row += 1
row += 1

banner(ws, row, "RESERVED — 4 REACTIVE STORY SLOTS (do not pre-schedule)"); row += 1
REACTIVE = [
    "The last heavy rain of the season, live — post the next day it genuinely pours.",
    "The villa is empty tonight — only post when it is.",
    "Snooker table / turf progress — when something actually changes on site. "
    "CONFIRM: is the turf still happening? The working sheet swapped it for the snooker table.",
    "A real WhatsApp inquiry, name blurred — 'this came in at 11pm. on a tuesday.'",
]
for r in REACTIVE:
    banner(ws, row, r, bg=WARM_WHITE, fg=BRAND_BROWN, bold=False, size=9)
    row += 1
row += 1

row += 1
banner(ws, row, "REFERENCE LINKS"); row += 1
for label, url in [("Instagram account", ACCOUNT),
                   ("August 2026 content calendar (source structure)", AUGUST_CC),
                   ("Bougain Villa Content machine (working sheet)", SOURCE_CC),
                   ("Font reference for the designer", LINKS["FONT REFERENCE"])]:
    style_cell(ws, "A" + str(row), label, f(9, bold=True), WARM_WHITE)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    c = style_cell(ws, "B" + str(row), url, f(9, color="1155CC"), WARM_WHITE)
    c.hyperlink = url
    for col in range(2, 8):
        ws.cell(row=row, column=col).border = BOX
    row += 1
row += 1

banner(ws, row, "CONFIRM BEFORE THIS SHIPS", bg=TERRACOTTA); row += 1
CONFIRM = [
    "1 · Ganesh Chaturthi, Anant Chaturdashi and Pitru Paksha exact dates, against a "
    "Maharashtra panchang. Planned here as Sep 14, Sep 24, Sep 26. Some panchangs give "
    "Chaturthi as Sep 15 — that moves Static 2, the silent week and the visarjan gap.",
    "2 · Regular price and weekday price, for the OFFERS highlight frames only.",
    "3 · WhatsApp number for the lockup.",
    "4 · Sleeps [N] — the capacity number. Still open from August.",
    "5 · Pet policy specifics — size limit, charge, which spaces.",
    "6 · Check-in and check-out times.",
    "7 · Three guest reviews, verbatim, for Story 16.",
    "8 · Confirm gym, library, snooker, TT and carrom are installed and shootable.",
    "9 · Is the turf still happening?",
    "10 · Exact brand brown and terracotta hex off the logo file. Still matched by eye.",
]
for r in CONFIRM:
    banner(ws, row, r, bg=WARM_WHITE, fg=BRAND_BROWN, bold=False, size=9)
    ws.row_dimensions[row].height = 24
    row += 1

for i in range(1, 8):
    ws.column_dimensions[get_column_letter(i)].width = 34
ws.freeze_panes = "A4"
ws.sheet_view.showGridLines = False



# --------------------------------------------------------------- helpers
def new_sheet(title, widths, title_text, sub_text=None):
    s = wb.create_sheet(title)
    for i, w in enumerate(widths, start=1):
        s.column_dimensions[get_column_letter(i)].width = w
    n = len(widths)
    banner(s, 1, title_text, span=n, size=13)
    s.row_dimensions[1].height = 28
    if sub_text:
        banner(s, 2, sub_text, span=n, size=9, bg=FILL_SUBHEAD,
               fg=BRAND_BROWN, bold=False)
        s.row_dimensions[2].height = 22
    s.sheet_view.showGridLines = False
    return s

def head_row(s, row, headers):
    for i, h in enumerate(headers, start=1):
        c = s.cell(row=row, column=i, value=h)
        c.font = f(9, bold=True, color=CREAM)
        c.fill = fill(BRAND_BROWN)
        c.alignment = align("left", "center")
        c.border = BOX
    s.row_dimensions[row].height = 22

def data_row(s, row, values, height=None, bg=None, size=9, bold_first=False):
    for i, v in enumerate(values, start=1):
        c = s.cell(row=row, column=i, value=v)
        c.font = f(size, bold=(bold_first and i == 1))
        c.fill = fill(bg or (WARM_WHITE if row % 2 else CREAM))
        c.alignment = align("left", "top")
        c.border = BOX
    if height:
        s.row_dimensions[row].height = height

def section(s, row, text, span, bg=FILL_SUBHEAD):
    banner(s, row, text, span=span, size=10, bg=bg, fg=BRAND_BROWN)
    return row + 1


# ==================================================================== SHEET 2
W = [26, 10, 40, 40, 40]
s2 = new_sheet("Designer Brief", W,
               "SEPTEMBER 2026  ·  DESIGNER BRIEF  ·  posts Sep 7–30",
               "Same structure as the August brief. Sections 1–4 are carried over "
               "unchanged unless marked SEPTEMBER. Section 5 is piece-by-piece; "
               "section 6 is new — highlight covers.")
r = 3

r = section(s2, r, "1 · CANVAS & EXPORT — every carousel slide and every static post", 5)
CANVAS = [
    ("Size", "1080 × 1350 px (4:5 portrait) — the tallest ratio Instagram allows in feed. Never square."),
    ("Grid-crop safe zone", "Instagram crops the PROFILE GRID thumbnail to a centre square. Keep the subject and all type inside the centre 1080 × 1080 — anything in the top or bottom 135 px is cut off on the grid."),
    ("Text margin", "100 px minimum from every edge. If it feels too empty, it is correct."),
    ("Export", "JPG · sRGB · quality 90 · under 8 MB per image"),
    ("Stories & highlights", "1080 × 1920 px (9:16). Keep all type inside the centre 1080 × 1420 — the top and bottom bands are covered by Instagram's own UI."),
    ("Never", "Do not send finals through WhatsApp or screenshot them. Re-compression destroys the gradients in the rain and lamp-light shots. Use Drive."),
]
for k, v in CANVAS:
    s2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    data_row(s2, r, [k, v], height=34, bold_first=True)
    r += 1
r += 1

r = section(s2, r, "2 · PALETTE — warm boho neutrals, taken from the live account", 5)
PALETTE = [
    ("Cream", CREAM, "Backgrounds, review cards, the CTA slide, highlight covers"),
    ("Warm white", WARM_WHITE, "Negative space, breathing room"),
    ("Sand", SAND, "Accents, dividers, secondary type"),
    ("Brand brown", BRAND_BROWN, "Primary type, logo, WhatsApp lockup, highlight cover labels"),
    ("Terracotta", TERRACOTTA, "One accent per piece, sparingly"),
]
for name, hexv, use in PALETTE:
    s2.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    data_row(s2, r, [name, "#" + hexv, use], height=18, bold_first=True)
    s2.cell(row=r, column=2).fill = fill(hexv)
    s2.cell(row=r, column=2).alignment = align("center", "center")
    r += 1
s2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
data_row(s2, r, ["CONFIRM", "Brand brown and terracotta are still matched by eye from the "
                 "Instagram grid — unresolved since August. Pull the exact values off the "
                 "logo file before the first export."], height=32, bg=FILL_FESTIVAL,
         bold_first=True)
r += 2

r = section(s2, r, "3 · TYPE, LOGO & CTA", 5)
TYPE = [
    ("Typeface", "ONE typeface across all eleven pieces AND all seven highlight covers. Either a refined serif or a light geometric sans — never both, never a third for accents."),
    ("Font reference", "A font reference was supplied this month (see Inspo Links sheet, 'Font for GD'). Pick ONE face from it. The one-typeface rule is unchanged — the reference is a starting point, not a licence for a second face."),
    ("Headline copy", "Lowercase. Small. Left-aligned in the top or bottom third. Never centred over a face, the pool, the lake, or the focal point of the photo."),
    ("Type colour", "Brand brown on light backgrounds, cream on dark. No pure black (#000) and no pure white (#FFF) anywhere — both read cheap against warm neutrals."),
    ("Logo", "Brushed-brown botanical mark, max 90 px wide. Cover slide only for carousels — not on every slide. Bottom-centre or top-left."),
    ("WhatsApp lockup", "Icon + number, bottom third, ~48 px tall, in BRAND BROWN. Not WhatsApp green — the green breaks the palette on every single piece."),
    ("Carousel ending", "Every carousel ends on a full WhatsApp slide — a whole cream card, not a corner overlay. This is the slide people screenshot."),
]
for k, v in TYPE:
    s2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    data_row(s2, r, [k, v], height=34, bold_first=True)
    r += 1
r += 1

r = section(s2, r, "4 · DO NOT — each has a specific reason, none are style preferences", 5)
DONT = [
    ("No magenta / pink", "The account's equity is warm neutrals. Bougainvillea pink was an early wrong assumption — confirmed wrong against the live profile."),
    ("No WhatsApp green", "Breaks the palette. Use the icon shape in brand brown."),
    ("No emoji or stickers", "No starbursts, no badges, no 'SWIPE →' arrows. The whole positioning is quiet luxury. The emoji in the highlight names are for the sheet only — they do not go on the artwork."),
    ("No shadows or glows", "No drop shadow, outer glow, or stroke on type. If type is unreadable, move it — do not add effects."),
    ("No dark gradient scrims", "Do not lay a black gradient over a photo to make text readable. Move the text into the sky, the water, or the wall instead."),
    ("No rupee figures in feed", "SEPTEMBER — the exception this month is the OFFERS highlight, and only there. Nothing in the feed carries a rate. The price question is what starts the WhatsApp conversation."),
    ("No urgency graphics", "No 'limited time', no countdowns, no 'only 2 left'. Weekday inventory is genuinely open and false scarcity gets noticed."),
    ("No non-veg or bar imagery from Sep 14", "SEPTEMBER — Shravan ended, but Ganpati runs Sep 14–24 and Pitru Paksha from Sep 26. Safe window for food or bar imagery is Sep 1–13 only."),
    ("No murti or visarjan imagery", "SEPTEMBER — a greeting, a diya, a marigold detail, the villa. Nothing devotional used as a marketing frame. No idol photography at all."),
    ("No offer creative Sep 14–30", "SEPTEMBER — the festival fortnight and Pitru Paksha are not selling windows. Pushing a discount into them reads badly and does not convert."),
]
for k, v in DONT:
    s2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    data_row(s2, r, [k, v], height=34, bold_first=True)
    r += 1
r += 1

r = section(s2, r, "5 · PIECE-BY-PIECE — eleven pieces, slide by slide", 5)
head_row(s2, r, ["PIECE", "SLIDE", "VISUAL", "ON-SCREEN COPY", "NOTES"])
r += 1

PIECES = [
    ("REEL 1 · All Of It · Sep 7", "—",
     "One continuous move through the house: exterior, living room, stairs, pool deck, lake. Held shots, no whip transitions.",
     "the whole house.\nin one go.",
     "Opens the month. Native audio, minimal music bed. Establishes every asset the rest of the month refers back to. 25–35 sec."),
    ("CAROUSEL 1 · What A Weekday Gets You · Sep 9", "1",
     "The pool, empty, morning light.", "what a weekday gets you.",
     "Cover. THE hardest-working piece of the month — moved to week one because the back half of September has no demand."),
    ("CAROUSEL 1 · What A Weekday Gets You · Sep 9", "2",
     "The pool again, wider, nobody in it.", "the pool. nobody in it.", ""),
    ("CAROUSEL 1 · What A Weekday Gets You · Sep 9", "3",
     "The lake from the deck, mid-morning.", "the lake. still there.",
     "New slide vs August — the lake is this month's headline asset."),
    ("CAROUSEL 1 · What A Weekday Gets You · Sep 9", "4",
     "The theatre, screen glow.", "the cinema. at 3pm.",
     "The differentiator. No competing Karjat villa has this."),
    ("CAROUSEL 1 · What A Weekday Gets You · Sep 9", "5",
     "Wide interior, empty, warm.", "the whole villa. just yours.", ""),
    ("CAROUSEL 1 · What A Weekday Gets You · Sep 9", "6",
     "Full cream card, type only.", "monday to thursday. 10% off.",
     "No rupee figure. None."),
    ("CAROUSEL 1 · What A Weekday Gets You · Sep 9", "7",
     "Full cream card.", "ask us for this week's dates.", "WhatsApp lockup."),
    ("STATIC 1 · The Lake Is The View · Sep 8", "—",
     "Lake-facing frame — the pool edge running into the lake, or the deck looking out. Huge negative space in the sky.",
     "the lake was here first.\nwe just pointed the house at it.",
     "The asset August never used. Type sits in the sky or the water, never over the horizon line. Number + WhatsApp icon bottom third, brand brown."),
    ("REEL 2 · The Pool, Day To Night · Sep 10", "—",
     "Same pool, four times of day, cut in order: morning cold light, six o'clock orange on the lake, nine o'clock with the lights on, and 2 PM on a weekday with nobody in it.",
     "one pool.\nfour times of day.",
     "The 2 PM weekday shot is the one that sells — hold it longest. Feeds the POOL highlight directly."),
    ("CAROUSEL 2 · Seven Things Inside · Sep 11", "1",
     "Wide interior, warm, looking through to the pool.", "seven reasons the car stays parked.",
     "Cover. Sells the experience, not the amenity list."),
    ("CAROUSEL 2 · Seven Things Inside · Sep 11", "2",
     "Infinity pool.", "one · the pool that doesn't end.", ""),
    ("CAROUSEL 2 · Seven Things Inside · Sep 11", "3",
     "Lake view from the deck.", "two · the lake.", ""),
    ("CAROUSEL 2 · Seven Things Inside · Sep 11", "4",
     "Home theatre, screen glow, seats.", "three · the cinema.", ""),
    ("CAROUSEL 2 · Seven Things Inside · Sep 11", "5",
     "Gym, and the games room — snooker, TT, carrom. Split or two frames.",
     "four · the gym.\nfive · snooker, tt, carrom.",
     "CONFIRM these are installed and shootable before this slide is designed."),
    ("CAROUSEL 2 · Seven Things Inside · Sep 11", "6",
     "The library corner, warm lamp.", "six · the library.",
     "The unexpected one. Nobody advertises a library — that is exactly why it gets saved."),
    ("CAROUSEL 2 · Seven Things Inside · Sep 11", "7",
     "Garden, early evening.", "seven · the garden.", ""),
    ("CAROUSEL 2 · Seven Things Inside · Sep 11", "8",
     "Full cream card.", "all of it, for one group.", "WhatsApp lockup."),
    ("STATIC 2 · Ganpati Greeting · Sep 14 · 9:00 AM", "—",
     "Warm frame — a diya, a marigold detail, or the villa entrance at first light. Family-suggestive without any people and without any idol.",
     "गणपती बाप्पा मोरया.",
     "GREETING ONLY. No offer, no CTA, no discount, no WhatsApp lockup, no clip-art, no murti, no modak graphics. The restraint is the point. Devanagari set in the same single typeface if it has the glyphs; otherwise pair with one matching Devanagari face and note it."),
    ("CAROUSEL 3 · A Guide To Karjat · Sep 16", "1",
     "Mist over the Sahyadris, wide. Type in the sky.", "a guide to karjat.",
     "Cover. BUILT FOR SAVES — make slide 1 look like a guide, not an ad. Saves compound during a high-attention festival week."),
    ("CAROUSEL 3 · A Guide To Karjat · Sep 16", "2",
     "A waterfall near Karjat, still running in September.", "the waterfalls are still running.",
     "Name the actual spot. Specifics are what get saved."),
    ("CAROUSEL 3 · A Guide To Karjat · Sep 16", "3",
     "The drive up — windscreen or roadside.", "the drive is half the point.", ""),
    ("CAROUSEL 3 · A Guide To Karjat · Sep 16", "4",
     "A named local spot worth going to.", "and this, if you have a morning.",
     "Name it. Vague slides do not get saved."),
    ("CAROUSEL 3 · A Guide To Karjat · Sep 16", "5",
     "The villa, lake in frame.", "or don't leave at all.",
     "The turn — guide becomes the villa."),
    ("CAROUSEL 3 · A Guide To Karjat · Sep 16", "6",
     "Full cream card.", "1.5 hrs from navi mumbai.", "WhatsApp lockup."),
    ("REEL 3 · After The City · Sep 21", "—",
     "Groups: friends around the snooker table, a table laid, the pool at night with people in it. Warm, populated, loud — the opposite of the quiet festival fortnight that preceded it.",
     "ten days of the city.\nnow this.",
     "The re-entry reel. First push since Sep 13. Feeds the GROUPS highlight."),
    ("CAROUSEL 4 · Who It's For · Sep 23", "1",
     "A group on the deck, lake behind.", "who it's for.", "Cover."),
    ("CAROUSEL 4 · Who It's For · Sep 23", "2",
     "Friends — games room or pool, evening.", "nine people who haven't been in one room since college.", ""),
    ("CAROUSEL 4 · Who It's For · Sep 23", "3",
     "Family — long table, kitchen or garden.", "a family that keeps saying next year.", ""),
    ("CAROUSEL 4 · Who It's For · Sep 23", "4",
     "A celebration setup — cake on the long table, or the garden lit at night. No balloons, no props.",
     "one birthday that deserves more than a restaurant.", ""),
    ("CAROUSEL 4 · Who It's For · Sep 23", "5",
     "Plan-view illustration of the villa on cream.", "four bedrooms. sleeps [CONFIRM N].",
     "Confirm the sleeps number. Still open from August."),
    ("CAROUSEL 4 · Who It's For · Sep 23", "6",
     "Full cream card.", "the whole villa goes to one group. always.",
     "This is the positioning line of the month. Never shared, never split."),
    ("CAROUSEL 4 · Who It's For · Sep 23", "7",
     "Full cream card.", "tell us the dates and how many.", "WhatsApp lockup."),
    ("REEL 4 · October Is Clear · Sep 28", "—",
     "Post-monsoon: everything still wet-season green, clear sky, the lake full and flat. No rain in a single frame.",
     "the rain stops.\nthe green stays.",
     "Pre-sells October — Navratri and Diwali. NO September offer anywhere in this. Pitru Paksha is running; this is a look-ahead, not a push."),
    ("REEL 5 · Book Now · Sep 30", "—",
     "Simple and short. The WhatsApp thread, the villa, the lake. Under 20 seconds.",
     "tell us the dates.\nand how many.",
     "Closes the month on conversion, pointed at October. Feeds the OFFERS highlight."),
]
for p in PIECES:
    data_row(s2, r, list(p), height=52)
    r += 1
r += 1

r = section(s2, r, "6 · HIGHLIGHT COVERS — seven covers, one system", 5)
COVERS = [
    ("Format", "1080 × 1920 px, designed so the centre circle crop (roughly the middle 640 × 640) carries the whole design. Anything outside that circle is invisible on the profile."),
    ("Ground", "Cream #F2EBE1 on all seven. Identical. The tray must read as one object, not seven."),
    ("Label", "Lowercase, brand brown #6B4F3A, centred, in the month's single typeface. One word each: offers · villa · experiences · rooms · pool · groups · karjat."),
    ("Mark", "Optional single botanical line-mark above the label, sand #C8A882, same scale on all seven. If it crowds the circle, drop it on all seven — never on some."),
    ("Do not", "No photographs on the covers. No emoji. No icons per category. No colour-coding between them. The emoji in this sheet are reference labels only."),
    ("Order in tray", "OFFERS · VILLA · EXPERIENCES · ROOMS · POOL · GROUPS · KARJAT — left to right. Instagram orders by most-recently-updated, so add a frame to them in reverse order on the day you set the tray."),
]
for k, v in COVERS:
    s2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    data_row(s2, r, [k, v], height=36, bold_first=True)
    r += 1
r += 1

s2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
data_row(s2, r, ["Copy in the ON-SCREEN COPY column is final and approved — set it verbatim, "
                 "including the lowercase. Square brackets mark the only things still to be "
                 "supplied."], height=30, bg=FILL_SUBHEAD)
s2.freeze_panes = "A3"
print("sheet 2 done, rows", r)


# ==================================================================== SHEET 3
s3 = new_sheet("Stories", [5, 11, 8, 10, 20, 30, 30, 20, 22],
               "SEPTEMBER 2026  ·  STORIES  ·  all 20, Sep 7–30",
               "Stories are the raw material. Highlights are assembled from them — which is "
               "why every highlight build date sits AFTER the stories that fill it. "
               "On-screen copy is set verbatim, lowercase, two lines split at the ' / '.")
head_row(s3, 3, ["#", "Date", "Day", "Time", "Story", "On-screen copy", "Visual",
                 "Sticker / CTA", "→ Highlight"])
STORIES = [
    (1, "Sep 7", "Mon", "9:30 AM", "VILLA", "all of it. / in ninety seconds.",
     "One continuous walk: gate, exterior, living room, stairs, pool deck, lake. No cuts if possible.",
     "—", "VILLA"),
    (2, "Sep 8", "Tue", "9:30 PM", "ROOMS", "four bedrooms. / no one gets the sofa.",
     "Each bedroom, warm lamp on, bed made. Vary the light between them so it doesn't read as one room.",
     "—", "ROOMS"),
    (3, "Sep 9", "Wed", "9:30 PM", "POOL", "the infinity pool. / it doesn't end.",
     "The edge where the water meets the lake. Hold the shot.", "—", "POOL"),
    (4, "Sep 10", "Thu", "9:30 AM", "LAKE VIEW", "the lake was here first. / we just pointed the house at it.",
     "From the deck, looking out. Morning haze if you can get it.", "—", "VILLA · POOL · KARJAT"),
    (5, "Sep 11", "Fri", "9:30 PM", "AMENITIES", "there is more inside / than you'd expect.",
     "Fast set: gym, theatre door, library shelf, games table. Four beats, no captions on each.",
     "—", "EXPERIENCES"),
    (6, "Sep 11", "Fri", "9:30 PM", "GAMES", "snooker. table tennis. carrom. / someone always loses.",
     "Hands on the table, mid-game. Movement, not a still room.",
     "Poll: snooker / tt / carrom", "EXPERIENCES"),
    (7, "Sep 12", "Sat", "9:30 AM", "THEATRE", "and there is a cinema. / inside the house.",
     "Screen glow, seats, lights low. Shot from the back row.", "—", "EXPERIENCES"),
    (8, "Sep 12", "Sat", "9:30 AM", "KITCHEN", "cook, or don't. / both are handled.",
     "Open kitchen, counter laid out, something actually on the hob.", "—", "EXPERIENCES"),
    (9, "Sep 13", "Sun", "9:30 AM", "OUTDOOR", "the garden at six. / the poolside at nine.",
     "Two shots, split screen or back to back — golden hour garden, lit poolside at night.",
     "—", "EXPERIENCES"),
    (10, "Sep 13", "Sun", "9:30 AM", "CAPACITY", "four bedrooms. sleeps [CONFIRM N]. / bring everyone.",
     "Plan-view graphic on cream, or the four doorways lit down a corridor.",
     "Poll: friends / family", "GROUPS"),
    (11, "Sep 17", "Thu", "9:30 AM", "LOCATION", "karjat. / the green starts before the exit.",
     "The approach road, then the Maps pin.", "Location sticker", "KARJAT"),
    (12, "Sep 18", "Fri", "9:30 PM", "HOW TO REACH", "1.5 hrs from navi mumbai. / 2 hrs from pune.",
     "Route graphic on cream with the two drive times and one landmark each. Confirm the Pune time.",
     "Question box", "KARJAT"),
    (13, "Sep 19", "Sat", "9:30 AM", "PET FRIENDLY", "your dog is invited. / the garden is theirs.",
     "A dog in the garden or on the deck. If there is no resident dog, shoot a guest's — do not stock-image this.",
     "Question box", "GROUPS"),
    (14, "Sep 21", "Mon", "9:30 AM", "STAY", "arrive at two. / stop checking your phone by four.",
     "Arrival beat: car doors, bags down, first drink on the deck.", "—", "VILLA"),
    (15, "Sep 22", "Tue", "9:30 AM", "CELEBRATIONS", "birthdays. anniversaries. / the whole house is yours.",
     "A laid long table, or the garden lit at night. No balloons, no props, no banners.",
     "—", "GROUPS"),
    (16, "Sep 23", "Wed", "9:30 PM", "REVIEWS", "[review 1 verbatim]",
     "Review text on a cream card, paired with a photo of the exact space it mentions. Three frames, one review each.",
     "—", "GROUPS"),
    (17, "Sep 25", "Fri", "9:30 AM", "CHECK-IN", "check in [CONFIRM]. / check out [CONFIRM].",
     "Cream card, type only. Keys, or the front door.", "—", "OFFERS"),
    (18, "Sep 26", "Sat", "9:30 AM", "RULES", "a few house rules. / all of them reasonable.",
     "Cream cards. Four rules maximum, one per frame. Noise, pool timings, pets, smoking.",
     "—", "OFFERS"),
    (19, "Sep 28", "Mon", "9:30 AM", "FAQs", "ask us anything. / we answer everything.",
     "Cream cards. The five questions that actually arrive on WhatsApp.",
     "Question box", "OFFERS"),
    (20, "Sep 29", "Tue", "9:30 AM", "BOOK NOW", "october is open. / whatsapp us.",
     "The WhatsApp thread on screen, name blurred, then the lake.",
     "Link / DM sticker", "OFFERS"),
]
r = 4
for st in STORIES:
    data_row(s3, r, list(st), height=46)
    r += 1
r += 1
s3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
data_row(s3, r, ["RESERVED — 4 reactive story slots, not pre-scheduled: last heavy rain of the "
                 "season · the villa is empty tonight · snooker table or turf progress · a real "
                 "WhatsApp inquiry with the name blurred."], height=28, bg=FILL_SUBHEAD)
s3.freeze_panes = "A4"
print("sheet 3 done")


# ==================================================================== SHEET 4
s4 = new_sheet("Highlights", [7, 18, 14, 44, 26, 12, 12, 34],
               "SEPTEMBER 2026  ·  HIGHLIGHTS  ·  all seven built by Sep 30",
               "Tray order is left to right on the profile. Note the inversion: OFFERS sits "
               "FIRST in the tray but is built LAST, because it is the only highlight that "
               "expires. The other six are evergreen — built once and left alone.")
head_row(s4, 3, ["Tray", "Highlight", "Cover label", "Frames, in order", "Fed by",
                 "Build date", "Rebuild?", "Notes"])
HL = [
    (1, "OFFERS 🎁", "offers",
     "1 · Monday–Thursday offer  2 · regular price → weekday price  3 · what's included  "
     "4 · validity  5 · CTA — 'DM / WhatsApp for available dates'",
     "Purpose-shot frames + Stories 17–20", "Sep 30", "MONTHLY",
     "THE ONLY HIGHLIGHT CARRYING PRICES. Struck-through regular → weekday, per your call. "
     "Built last, pinned first. Must be rebuilt whenever the offer or validity changes — "
     "an expired offer sitting at the front of the tray is worse than no offer highlight."),
    (2, "VILLA 🏡", "villa",
     "1 · Exterior  2 · Living room  3 · Bedrooms  4 · Pool  5 · Lake view  6 · Night view",
     "Stories 1, 4, 14", "Sep 13", "No",
     "The 5–6 frame overview. This is the one a cold visitor opens first — it has to answer "
     "'what is this place' without a single word of sell."),
    (3, "EXPERIENCES ✨", "experiences",
     "1 · Infinity pool  2 · Lake view  3 · Home theatre  4 · Gym  5 · Snooker / TT / carrom  "
     "6 · Library  7 · Garden",
     "Stories 5, 6, 7, 8, 9", "Sep 13", "No",
     "Sells the experience, not the amenity list. The library is the sleeper — nobody "
     "advertises one, which is exactly why it lands."),
    (4, "ROOMS 🛏️", "rooms",
     "All four bedrooms, then bathrooms, then the bathtub",
     "Story 2", "Sep 13", "No",
     "Vary the light between bedrooms so it does not read as one room shot four times. "
     "The bathtub is a booking driver — give it its own frame, not a corner of a bathroom shot."),
    (5, "POOL 🌊", "pool",
     "1 · Day  2 · Sunset  3 · Night  4 · Lake view from the water",
     "Stories 3, 4 + Reel 2 frames", "Sep 13", "No",
     "Make this visually strong — it is the most-tapped highlight on almost every villa "
     "account. Pull the four beats straight from Reel 2."),
    (6, "GROUPS 👥", "groups",
     "1 · Friends  2 · Families  3 · Celebrations  4 · Maximum capacity  "
     "5 · Private villa — never shared",
     "Stories 10, 13, 15, 16", "Sep 27", "No",
     "Your core market. Frame 5 is the positioning line of the month: the whole villa goes to "
     "one group, always. Never shared, never split."),
    (7, "KARJAT 🌿", "karjat",
     "1 · Karjat scenery  2 · Nearby attractions  3 · The drive and route  4 · Escape the city",
     "Stories 11, 12", "Sep 20", "No",
     "Sells the escape, not the property. This is also where the drive times live, so it "
     "absorbs the most common pre-booking question."),
]
r = 4
for h in HL:
    data_row(s4, r, list(h), height=68)
    r += 1
r += 2

r = section(s4, r, "BUILD SEQUENCE — in date order", 8)
head_row(s4, r, ["", "Date", "Build", "Why then", "", "", "", ""])
r += 1
SEQ = [
    ("", "Sep 13", "VILLA · ROOMS · POOL · EXPERIENCES",
     "Stories 1–10 all ran Sep 7–13. Ganpati eve is a quiet posting day, so it is free build time. Four highlights go up at once.", "", "", "", ""),
    ("", "Sep 20", "KARJAT",
     "Stories 11–12 ran Sep 17–18. Quiet visarjan Sunday.", "", "", "", ""),
    ("", "Sep 27", "GROUPS",
     "Stories 10, 13, 15, 16 are all in by Sep 23.", "", "", "", ""),
    ("", "Sep 30", "OFFERS  +  SET TRAY ORDER",
     "Built last so prices and validity are current on the day the tray goes live. Needs the two price figures — see CONFIRM. Set the tray order the same day.", "", "", "", ""),
]
for q in SEQ:
    data_row(s4, r, list(q), height=32)
    r += 1
r += 1
s4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
data_row(s4, r, ["Instagram orders the highlight tray by most-recently-updated, not by "
                 "creation date. To land the OFFERS · VILLA · EXPERIENCES · ROOMS · POOL · "
                 "GROUPS · KARJAT order, add one frame to each highlight in REVERSE order "
                 "(karjat first, offers last) on Sep 30."],
         height=32, bg=FILL_FESTIVAL)
s4.freeze_panes = "A4"
print("sheet 4 done")


# ==================================================================== SHEET 5
s5 = new_sheet("Captions", [6, 24, 10, 9, 62, 46, 22, 40],
               "SEPTEMBER 2026  ·  CAPTIONS  ·  all 11 feed pieces, Sep 7–30",
               "August voice throughout: lowercase, short lines, no exclamation marks, no "
               "emoji, no rupee figures, CTA always to WhatsApp. Line breaks are intentional — "
               "paste with them intact. Hashtags go in the caption, not a first comment: "
               "Instagram weights them identically and a first comment can be missed.")
head_row(s5, 3, ["#", "Piece", "Date", "Time", "Caption", "Hashtags", "Notes", "Inspo link"])

CAPS = [
    (1, "REEL 1 · All Of It", "Sep 7", "9:30 PM",
     "the whole house, in one go.\n\n"
     "four bedrooms. an infinity pool pointed at a lake.\n"
     "a cinema inside. a gym. a library nobody expects.\n\n"
     "karjat. an hour and a half from navi mumbai.\n\n"
     "monday to thursday is 10% off.\n"
     "whatsapp us and we'll tell you what's open.",
     "#bougainvilla #karjat #karjatvilla #villainkarjat #karjatfarmhouse #privatevilla "
     "#lakeviewvilla #villawithprivatepool #infinitypool #weekendgetaway "
     "#mumbaiweekendgetaway #punegetaway #4bhkvilla #luxuryvillaindia #boutiquevilla",
     "Opens the month. Brand-first, offer mentioned once at the end."),
    (2, "CAROUSEL 1 · What A Weekday Gets You", "Sep 9", "9:30 AM",
     "the pool, with nobody in it.\n\n"
     "that is the actual difference between a saturday here\n"
     "and a tuesday.\n\n"
     "same four bedrooms. same lake. same cinema.\n"
     "same house. 10% off, monday to thursday.\n\n"
     "we don't publish rates — ask us.\n"
     "it takes one message.",
     "#karjatvilla #villainkarjat #weekdaygetaway #karjat #privatevilla #lakeviewvilla "
     "#villawithprivatepool #weekendgetaway #mumbaiweekendgetaway #punegetaway "
     "#bougainvilla #4bhkvilla #luxuryvilla #maharashtratourism",
     "The offer post. No rupee figure anywhere in the caption. The 'we don't publish rates' "
     "line is doing the conversion work — it makes the DM the only way to find out."),
    (3, "STATIC 1 · The Lake Is The View", "Sep 8", "9:30 AM",
     "the lake was here first.\n"
     "we just pointed the house at it.\n\n"
     "every room on that side looks at water.\n\n"
     "karjat. whatsapp us for dates.",
     "#lakeviewvilla #karjat #karjatvilla #villainkarjat #lakeview #privatevilla "
     "#bougainvilla #weekendgetaway #mumbaiweekendgetaway #sahyadri #boutiquevilla "
     "#villawithaview",
     "Brand anchor. Short caption — the frame carries it."),
    (4, "REEL 2 · The Pool, Day To Night", "Sep 10", "9:30 PM",
     "one pool. four times of day.\n\n"
     "morning, when it's still cold.\n"
     "six, when the lake goes orange.\n"
     "nine, when the lights come on.\n\n"
     "and the one nobody photographs —\n"
     "two in the afternoon, on a tuesday,\n"
     "with no one else in it.\n\n"
     "whatsapp us for weekday dates.",
     "#infinitypool #karjatvilla #villawithprivatepool #karjat #lakeviewvilla "
     "#villainkarjat #poolvilla #weekendgetaway #mumbaiweekendgetaway #punegetaway "
     "#bougainvilla #luxuryvillaindia",
     "The fourth beat is the sell. Everything before it is setup."),
    (5, "CAROUSEL 2 · Seven Things Inside", "Sep 11", "9:30 AM",
     "seven reasons the car stays parked.\n\n"
     "the infinity pool. the lake.\n"
     "the home theatre. the gym.\n"
     "snooker, table tennis, carrom.\n"
     "the library. the garden.\n\n"
     "you can do the whole weekend\n"
     "without leaving the gate.\n\n"
     "save this one. whatsapp us for dates.",
     "#karjatvilla #villainkarjat #karjat #privatevilla #hometheatre #villawithpool "
     "#lakeviewvilla #weekendgetaway #mumbaiweekendgetaway #punegetaway #bougainvilla "
     "#luxuryvilla #boutiquevilla #villaamenities",
     "Built for saves. 'save this one' is an explicit ask — the only place this month we use one."),
    (6, "STATIC 2 · Ganpati Greeting", "Sep 14", "9:00 AM",
     "गणपती बाप्पा मोरया.\n\n"
     "may this one be spent\n"
     "with everyone you meant to call.\n\n"
     "— from all of us at bougainvilla",
     "#ganeshchaturthi #ganpatibappamorya #bougainvilla #karjat",
     "NO offer. NO CTA. NO WhatsApp number. Four hashtags maximum — a long tag block on a "
     "festival greeting is exactly what makes it read as marketing. Do not boost this post."),
    (7, "CAROUSEL 3 · A Guide To Karjat", "Sep 16", "9:30 AM",
     "save this one.\n\n"
     "everything worth doing in karjat,\n"
     "in one place —\n\n"
     "the waterfalls, still running in september.\n"
     "the drive, which is half the point.\n"
     "the spots people actually go to,\n"
     "not the ones on the first page of google.\n\n"
     "and when you're done,\n"
     "somewhere to sleep with a lake in front of it.\n\n"
     "1.5 hrs from navi mumbai. 2 from pune.",
     "#karjat #karjattourism #karjatwaterfalls #maharashtratourism #sahyadri "
     "#weekendgetaway #mumbaiweekendgetaway #punegetaway #monsoongetaway #thingstodo "
     "#karjatvilla #villainkarjat #bougainvilla #traveldiaries",
     "The saves play. Post during Ganpati precisely because attention is high and booking "
     "intent is low — a save now converts in October. Name real places on the slides."),
    (8, "REEL 3 · After The City", "Sep 21", "9:30 PM",
     "ten days of the city. now this.\n\n"
     "the whole villa, one group, nobody else in it.\n"
     "four bedrooms, so nobody argues.\n"
     "a pool, a cinema, a snooker table.\n\n"
     "the october weekends are open.\n"
     "whatsapp us.",
     "#karjatvilla #villainkarjat #karjat #privatevilla #groupgetaway #friendsgetaway "
     "#weekendgetaway #mumbaiweekendgetaway #punegetaway #bougainvilla #4bhkvilla "
     "#lakeviewvilla #villawithprivatepool",
     "First push since Sep 13. Points at October, not the rest of September — Pitru Paksha "
     "starts in five days."),
    (9, "CAROUSEL 4 · Who It's For", "Sep 23", "9:30 AM",
     "nine people who haven't been\n"
     "in the same room since college.\n\n"
     "a family that keeps saying next year.\n\n"
     "one birthday that deserves\n"
     "more than a restaurant.\n\n"
     "the whole villa goes to one group. always.\n"
     "never shared, never split.\n\n"
     "tell us the dates and how many.",
     "#karjatvilla #villainkarjat #privatevilla #groupgetaway #familygetaway "
     "#birthdaycelebration #karjat #weekendgetaway #mumbaiweekendgetaway #punegetaway "
     "#bougainvilla #4bhkvilla #petfriendlyvilla #celebrationvenue",
     "'never shared, never split' is the positioning line of the month. Do not soften it."),
    (10, "REEL 4 · October Is Clear", "Sep 28", "9:30 PM",
     "the rain stops. the green stays.\n\n"
     "october in karjat is the version\n"
     "people don't know about —\n"
     "everything still monsoon green,\n"
     "none of the monsoon.\n\n"
     "navratri, diwali,\n"
     "and the four weekends in between.\n\n"
     "whatsapp us for dates.",
     "#karjat #karjatvilla #villainkarjat #octoberweekend #postmonsoon #navratri #diwali "
     "#weekendgetaway #mumbaiweekendgetaway #punegetaway #bougainvilla #lakeviewvilla "
     "#maharashtratourism",
     "Pure October pre-sell. No September offer mentioned — Pitru Paksha is running. "
     "No scarcity language: the weekends are genuinely open."),
    (11, "REEL 5 · Book Now", "Sep 30", "9:30 PM",
     "three ways. one answer.\n\n"
     "whatsapp — fastest, we reply the same day.\n"
     "call — if you'd rather talk it through.\n"
     "dm — this works too.\n\n"
     "tell us the dates and how many.\n"
     "that's the whole process.\n\n"
     "karjat. four bedrooms. a lake.",
     "#karjatvilla #villainkarjat #karjat #privatevilla #bookdirect #weekendgetaway "
     "#mumbaiweekendgetaway #punegetaway #bougainvilla #lakeviewvilla #4bhkvilla "
     "#villawithprivatepool",
     "Closes the month. Feeds the OFFERS highlight the same day."),
]
r = 4
for c in CAPS:
    url = LINKS.get(c[1], "")
    data_row(s5, r, list(c) + [url], height=150)
    if url:
        lc = s5.cell(row=r, column=8); lc.hyperlink = url; lc.font = f(9, color="1155CC")
    r += 1
r += 1
s5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
data_row(s5, r, ["No caption this month contains a rupee figure, an exclamation mark, an emoji, "
                 "or the words 'limited', 'hurry' or 'book now before'. That is deliberate and "
                 "carried from the August rules — please keep it that way through review."],
         height=30, bg=FILL_SUBHEAD)
s5.freeze_panes = "A4"
print("sheet 5 done")


# ==================================================================== SHEET 6
s6 = new_sheet("Inspo Links", [8, 30, 10, 62, 44, 20],
               "SEPTEMBER 2026  ·  INSPO LINKS  ·  reference only",
               "NOT REVIEWED. instagram.com is blocked by this session's network egress proxy, "
               "so none of these posts could be opened or assessed. They are carried through "
               "exactly as supplied. Designer: open each one yourself before you start.")
head_row(s6, 3, ["Ref", "Maps to piece", "Date", "Link", "What to take from it", "Reviewed?"])
INSPO = [
    ("Carousel 1", "CAROUSEL 1 · What A Weekday Gets You", "Sep 9",
     "https://www.instagram.com/p/DbsuoOVjPzV/",
     "Structure and slide pacing for the offer carousel. Client note carried from August: "
     "minimum 4 photos, WhatsApp on the very last slide.", "NOT REVIEWED"),
    ("Carousel 2", "CAROUSEL 2 · Seven Things Inside", "Sep 11",
     "https://www.instagram.com/p/DcVqkH6CP9E/",
     "Reference for the numbered-list carousel format — seven items without it reading as a "
     "spec sheet.", "NOT REVIEWED"),
    ("Carousel 3", "CAROUSEL 3 · A Guide To Karjat", "Sep 16",
     "https://www.instagram.com/p/Db71wKNAOQr/",
     "Guide-style cover treatment. Slide 1 must look like a guide, not an ad — that is what "
     "drives the save.", "NOT REVIEWED"),
    ("Carousel 4", "CAROUSEL 4 · Who It's For", "Sep 23",
     "https://www.instagram.com/p/DcX2SH-iByn/",
     "Reference for the audience/segment carousel — how to show three group types without "
     "captioning each one to death.", "NOT REVIEWED"),
    ("Static 1", "STATIC 1 · The Lake Is The View", "Sep 8",
     "https://www.instagram.com/p/DcQEepBDakg/?img_index=2",
     "Single-frame composition with heavy negative space. Note: the link points at image 2 of "
     "the set — that specific frame is the reference.", "NOT REVIEWED"),
    ("Static 2", "STATIC 2 · Ganpati Greeting", "Sep 14",
     "https://www.instagram.com/p/DbA4f-sNjtf/",
     "Restraint reference for the festival greeting. Greeting large, everything else small or "
     "absent.", "NOT REVIEWED"),
    ("Font for GD", "ALL PIECES — typeface selection", "—",
     "https://www.instagram.com/p/Dblrpj-E_Lf/?img_index=1&igsi=cWd0amt3MXdqZnAw",
     "Type reference. Pick ONE face from this. The one-typeface rule holds across all 11 "
     "pieces and all 7 highlight covers.", "NOT REVIEWED"),
    ("Reel 1", "REEL 1 · All Of It", "Sep 7",
     "https://www.instagram.com/p/DadUztOMe56/",
     "Pacing reference for the full-villa walkthrough.", "NOT REVIEWED"),
    ("Reel 2", "REEL 2 · The Pool, Day To Night", "Sep 10",
     "https://www.instagram.com/p/DceNoWbyQnA/",
     "Time-of-day progression cut.", "NOT REVIEWED"),
    ("Reel 3", "REEL 3 · After The City", "Sep 21",
     "https://www.instagram.com/p/DbvH3GeJSR0/",
     "Group/people-in-the-space energy.", "NOT REVIEWED"),
    ("Reel 4", "REEL 4 · October Is Clear", "Sep 28",
     "https://www.instagram.com/p/DcYxp75p8GM/",
     "Landscape and seasonal-change reference.", "NOT REVIEWED"),
    ("Reel 5", "REEL 5 · Book Now", "Sep 30",
     "https://www.instagram.com/p/DcOTWy6ML1t/?igsi=MWtjOHB0NnIybzgwYw==",
     "Fifth link from the source sheet, supplied unlabelled. Assigned here to the closing "
     "conversion reel — reassign if it was meant for something else.", "NOT REVIEWED"),
]
r = 4
for i in INSPO:
    data_row(s6, r, list(i), height=44)
    s6.cell(row=r, column=6).font = f(9, bold=True, color=TERRACOTTA)
    lc = s6.cell(row=r, column=4); lc.hyperlink = i[3]; lc.font = f(9, color="1155CC")
    r += 1
r += 1
s6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
data_row(s6, r, ["Account: " + ACCOUNT + "     August 2026 calendar: " + AUGUST_CC],
         height=26, bg=FILL_SUBHEAD)
s6.freeze_panes = "A4"
print("sheet 6 done")


# ==================================================================== SHEET 7
s7 = new_sheet("Production Tracker",
               [5, 34, 11, 10, 9, 13, 13, 13, 13, 15, 13, 17, 30, 40],
               "SEPTEMBER 2026  ·  PRODUCTION TRACKER  ·  posts Sep 7–30",
               "Columns match the 'Bougain Villa Content machine' sheet so this drops straight "
               "in. Status cells take one of: Not started / In progress / Done. The summary "
               "below counts them live — change a status and the counts update.")
r = 3
r = section(s7, r, "STATUS SUMMARY — live counts", 14)
head_row(s7, r, ["", "Stage", "Not started", "In progress", "Done", "Total items",
                 "", "", "", "", "", "", "", ""])
summary_head = r
r += 1

DATA_FIRST = 0  # filled in after the table is written
summary_rows = [("Shoot", "G"), ("Design", "H"), ("Caption", "I"), ("Scheduled", "K")]
summary_first = r
r += len(summary_rows)
r += 1

r = section(s7, r, "TRACKER", 14)
head_row(s7, r, ["#", "Item", "Type", "Post date", "Time", "Owner", "Shoot", "Design",
                 "Caption", "Deliverable sent", "Scheduled", "Revision review", "Comment",
                 "Inspo link"])
r += 1
DATA_FIRST = r

FEED = [
    ("REEL 1 · All Of It", "Reel", "Sep 7", "9:30 PM"),
    ("CAROUSEL 1 · What A Weekday Gets You", "Carousel", "Sep 9", "9:30 AM"),
    ("STATIC 1 · The Lake Is The View", "Static", "Sep 8", "9:30 AM"),
    ("REEL 2 · The Pool, Day To Night", "Reel", "Sep 10", "9:30 PM"),
    ("CAROUSEL 2 · Seven Things Inside", "Carousel", "Sep 11", "9:30 AM"),
    ("STATIC 2 · Ganpati Greeting", "Static", "Sep 14", "9:00 AM"),
    ("CAROUSEL 3 · A Guide To Karjat", "Carousel", "Sep 16", "9:30 AM"),
    ("REEL 3 · After The City", "Reel", "Sep 21", "9:30 PM"),
    ("CAROUSEL 4 · Who It's For", "Carousel", "Sep 23", "9:30 AM"),
    ("REEL 4 · October Is Clear", "Reel", "Sep 28", "9:30 PM"),
    ("REEL 5 · Book Now", "Reel", "Sep 30", "9:30 PM"),
]
STORY_ROWS = [(f"STORY {n} · {t}", "Story", d, tm)
              for n, d, _dy, tm, t, *_ in STORIES]
HL_ROWS = [("HIGHLIGHT · " + h[1].split(" ")[0], "Highlight", h[5], "—") for h in HL]

NOTES = {
    "STATIC 2 · Ganpati Greeting": "Confirm Chaturthi date before design. No offer, no CTA.",
    "CAROUSEL 4 · Who It's For": "Blocked on sleeps [N].",
    "STORY 10 · CAPACITY": "Blocked on sleeps [N].",
    "STORY 13 · PET FRIENDLY": "Blocked on pet policy.",
    "STORY 16 · REVIEWS": "Blocked on 3 verbatim reviews.",
    "STORY 17 · CHECK-IN": "Blocked on check-in / check-out times.",
    "HIGHLIGHT · OFFERS": "Blocked on regular + weekday price. Build LAST, pin FIRST.",
}

n = 1
for label, kind, date, tm in FEED + STORY_ROWS + HL_ROWS:
    url = LINKS.get(label, "")
    data_row(s7, r, [n, label, kind, date, tm, "[assign]", "Not started", "Not started",
                     "Not started", "", "Not started", "", NOTES.get(label, ""), url],
             height=18)
    if label in NOTES:
        s7.cell(row=r, column=13).font = f(9, bold=True, color=TERRACOTTA)
    if url:
        lc = s7.cell(row=r, column=14); lc.hyperlink = url; lc.font = f(9, color="1155CC")
    r += 1
    n += 1
DATA_LAST = r - 1

# summary formulas — written now that the table's extent is known
for i, (stage, col) in enumerate(summary_rows):
    rr = summary_first + i
    rng = f"{col}${DATA_FIRST}:{col}${DATA_LAST}"
    data_row(s7, rr, ["", stage,
                      f'=COUNTIF({rng},"Not started")',
                      f'=COUNTIF({rng},"In progress")',
                      f'=COUNTIF({rng},"Done")',
                      f'=COUNTA({rng})', "", "", "", "", "", "", "", ""], height=18)
    for c in range(3, 7):
        s7.cell(row=rr, column=c).alignment = align("center", "center")

r = DATA_LAST + 2
s7.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
data_row(s7, r, ["LEGEND — edit only the Owner, Shoot, Design, Caption, Deliverable sent, "
                 "Scheduled, Revision review and Comment columns. Status cells take exactly "
                 "'Not started', 'In progress' or 'Done' — any other spelling will not be "
                 "counted by the summary above."], height=30, bg=FILL_SUBHEAD)
r += 1
s7.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
data_row(s7, r, ["EXAMPLE OF EXPECTED FORMAT — delete this row before use: "
                 "1 | REEL 1 · All Of It | Reel | Sep 1 | 9:30 PM | Hardik | Done | "
                 "In progress | Done | 28 Aug | Not started | v2 sent, awaiting sign-off | "
                 "colour grade approved"], height=28, bg=WARM_WHITE)
s7.freeze_panes = "A" + str(DATA_FIRST)
print(f"sheet 7 done, data rows {DATA_FIRST}-{DATA_LAST}")

wb.save(OUT)
print("SAVED", OUT, "| sheets:", wb.sheetnames)
